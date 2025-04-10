# import xlrd
# import os
#
# import pandas as pd
#
#
# file_path = 'D:/My_Files/桌面/杭州港华/杭州字典表0208.xlsx'
# file_path_out = 'D:/My_Files/桌面/杭州港华/杭州字典表_1.xlsx'
#
# import openpyxl
# from openpyxl import Workbook
#
# # 打开原始Excel文件
# original_workbook = openpyxl.load_workbook(file_path )
#
# # 创建一个新的Excel工作簿
# new_workbook = Workbook()
# new_sheet = new_workbook.active
#
# # 获取原始Excel文件中的所有工作表名称
# sheet_names = original_workbook.sheetnames
#
# # 将工作表名称写入新的Excel文件
# for index, name in enumerate(sheet_names, start=1):
#     new_sheet.cell(row=index, column=1, value=name)
#
# # 保存新的Excel文件
# new_workbook.save(file_path_out)
#
# # 关闭工作簿
# original_workbook.close()
# new_workbook.close()



import openpyxl
from openpyxl import Workbook

file_path = 'D:/My_Files/桌面/YX20250305字典.xlsx'
file_path_out = 'D:/My_Files/桌面/YX20250305NEW.xlsx'

# 打开原始Excel文件
original_workbook = openpyxl.load_workbook(file_path)

# 创建一个新的Excel工作簿
new_workbook = Workbook()
new_sheet = new_workbook.active

# 获取原始Excel文件中的所有工作表名称
sheet_names = original_workbook.sheetnames

# 将工作表名称和B2单元格的值写入新的Excel文件
for index, name in enumerate(sheet_names, start=1):
    sheet = original_workbook[name]
    b2_value = sheet['B2'].value
    new_sheet.cell(row=index, column=1, value=name)
    new_sheet.cell(row=index, column=2, value=b2_value)

# 保存新的Excel文件
new_workbook.save(file_path_out)

# 关闭工作簿
original_workbook.close()
new_workbook.close()
